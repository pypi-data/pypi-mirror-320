import unittest
import tempfile
import shutil
import os
import json
import csv
import yaml
import pickle
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from unittest.mock import patch

# If your DataSaver class is in the same file, omit this import.
# Otherwise, ensure data_saver.py is in the same directory (or installed as a package).
from reelscraper.utils import DataSaver


class TestDataSaver(unittest.TestCase):
    """
    A unittest-based test suite for DataSaver class.
    """

    def setUp(self):
        """
        Create a temporary directory for each test to avoid polluting the filesystem.
        """
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        """
        Remove the temporary directory after each test.
        """
        shutil.rmtree(self.test_dir, ignore_errors=True)

    # -------------------------------------------------------------------------
    # Initialization Tests
    # -------------------------------------------------------------------------

    def test_init_invalid_folder_empty(self):
        """Test that an empty folder path raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder="", file_extension="json")

    def test_init_invalid_folder_not_string(self):
        """Test that a non-string folder raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder=None, file_extension="json")

    def test_init_invalid_file_extension_empty(self):
        """Test that an empty file extension raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder=self.test_dir, file_extension="")

    def test_init_invalid_file_extension_not_string(self):
        """Test that a non-string file extension raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder=self.test_dir, file_extension=None)

    def test_init_file_extension_with_dot_only(self):
        """Test that a file extension of '.' or similar raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder=self.test_dir, file_extension=".")

    def test_init_filename_with_extension_in_name(self):
        """Test that providing a file name containing '.' raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(
                folder=self.test_dir, file_extension="json", file_name="data.json"
            )

    def test_init_invalid_file_name(self):
        """Test that an empty or non-string file name raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder=self.test_dir, file_extension="json", file_name="")

    def test_init_unsupported_extension(self):
        """Test that an unsupported file extension raises a ValueError."""
        with self.assertRaises(ValueError):
            DataSaver(folder=self.test_dir, file_extension="unsupported_ext")

    def test_init_valid(self):
        """
        Test valid initialization does not raise an error.
        Also checks that the folder is created and the full path is correct.
        """
        ds = None
        try:
            ds = DataSaver(
                folder=self.test_dir, file_extension="json", file_name="valid_data"
            )
        except Exception as e:
            self.fail(f"Initialization raised an unexpected exception: {e}")

        self.assertIsNotNone(ds, "DataSaver instance should be created.")
        self.assertTrue(
            os.path.exists(self.test_dir), "Folder should exist after initialization."
        )
        expected_path = os.path.join(self.test_dir, "valid_data.json")
        self.assertEqual(
            ds.full_path, expected_path, "Full path should match expected."
        )

    def test_init_replace_file_overwrites(self):
        """
        Test that if replace_file=True and a file with the same path exists,
        the DataSaver still uses the same path without renaming.
        """
        file_path = os.path.join(self.test_dir, "my_data.json")

        # Create an existing file
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("Existing content")

        ds = DataSaver(
            folder=self.test_dir,
            file_extension="json",
            replace_file=True,
            file_name="my_data",
        )
        self.assertEqual(ds.full_path, file_path)
        self.assertTrue(os.path.exists(file_path))

    def test_init_no_replace_file_renames(self):
        """
        Test that if replace_file=False and a file with the same path exists,
        the DataSaver renames the file to avoid overwriting the existing one.
        """
        file_path = os.path.join(self.test_dir, "my_data.json")

        # Create an existing file
        with open(file_path, "w", encoding="utf-8") as f:
            f.write("Existing content")

        ds = DataSaver(
            folder=self.test_dir,
            file_extension="json",
            replace_file=False,
            file_name="my_data",
        )

        # The new file name should contain "_1" (or possibly "_2", etc.)
        self.assertNotEqual(
            ds.full_path, file_path, "Path should have been changed for a new file."
        )
        self.assertTrue(ds.full_path.endswith(".json"))
        self.assertIn("_1.json", ds.full_path)

    # -------------------------------------------------------------------------
    # Save Method Tests
    # -------------------------------------------------------------------------

    def test_save_empty_data_raises_error(self):
        """Test that saving empty data raises a ValueError."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_empty_data"
        )
        with self.assertRaises(ValueError):
            ds.save([])

    def test_save_non_list_raises_error(self):
        """Test that saving non-list raises a ValueError."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_non_list"
        )
        with self.assertRaises(ValueError):
            ds.save("not a list")

    def test_save_list_with_non_dict_items_raises_error(self):
        """Test that a list containing non-dict items raises a ValueError."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_mixed_list"
        )
        data = [{"valid": "dict"}, ["not", "a", "dict"]]
        with self.assertRaises(ValueError):
            ds.save(data)

    # -------------------------------------------------------------------------
    # File Format Specific Tests
    # -------------------------------------------------------------------------

    def test_save_json(self):
        """Test saving JSON data and verify file content."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_json"
        )
        data = [{"key": "value", "number": 123}, {"another_key": "another_value"}]
        ds.save(data)

        with open(ds.full_path, "r", encoding="utf-8") as f:
            loaded_data = json.load(f)

        self.assertEqual(
            data, loaded_data, "JSON data on disk should match the data saved."
        )

    def test_save_csv(self):
        """Test saving CSV data and verify file content."""
        ds = DataSaver(folder=self.test_dir, file_extension="csv", file_name="test_csv")
        data = [{"name": "John", "age": 30}, {"name": "Jane", "age": 25}]
        ds.save(data)

        rows = []
        with open(ds.full_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)

        # CSV DictReader loads everything as strings
        self.assertEqual(len(rows), 2, "There should be 2 rows in CSV.")
        self.assertEqual(rows[0]["name"], "John")
        self.assertEqual(rows[0]["age"], "30")
        self.assertEqual(rows[1]["name"], "Jane")
        self.assertEqual(rows[1]["age"], "25")

    def test_save_yaml(self):
        """Test saving YAML data and verify file content."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="yaml", file_name="test_yaml"
        )
        data = [{"env": "prod", "version": 1}, {"env": "dev", "version": 2}]
        ds.save(data)

        with open(ds.full_path, "r", encoding="utf-8") as f:
            loaded_data = yaml.safe_load(f)

        self.assertEqual(
            data, loaded_data, "YAML data on disk should match the data saved."
        )

    def test_save_xml(self):
        """Test saving XML data and verify file content."""
        ds = DataSaver(folder=self.test_dir, file_extension="xml", file_name="test_xml")
        data = [{"tag1": "value1", "tag2": "value2"}, {"tag3": "value3"}]
        ds.save(data)

        tree = ET.parse(ds.full_path)
        root = tree.getroot()

        # Expecting structure: <root><record><tag1>value1</tag1><tag2>value2</tag2>...</record>...</root>
        records = list(root.findall("record"))
        self.assertEqual(
            len(records), 2, "There should be 2 <record> elements in the XML root."
        )

        # Check first record
        first_record_tags = {child.tag: child.text for child in records[0]}
        self.assertEqual(first_record_tags["tag1"], "value1")
        self.assertEqual(first_record_tags["tag2"], "value2")

        # Check second record
        second_record_tags = {child.tag: child.text for child in records[1]}
        self.assertEqual(second_record_tags["tag3"], "value3")

    def test_save_pickle(self):
        """Test saving pickle data and verify file content."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="pkl", file_name="test_pickle"
        )
        data = [{"hello": "world"}, {"foo": "bar"}]
        ds.save(data)

        with open(ds.full_path, "rb") as f:
            loaded_data = pickle.load(f)

        self.assertEqual(
            data, loaded_data, "Pickle data on disk should match the data saved."
        )

    def test_save_xlsx(self):
        """Test saving XLSX data and verify file content."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="xlsx", file_name="test_xlsx"
        )
        data = [{"col1": "val1", "col2": 10}, {"col1": "val2", "col3": "something"}]
        ds.save(data)

        wb = load_workbook(ds.full_path)
        ws = wb.active

        # Header row
        headers = [cell.value for cell in ws[1]]
        # We expect columns col1, col2, col3 (though col3 is absent in first row)
        self.assertTrue("col1" in headers)
        self.assertTrue("col2" in headers or "col3" in headers)

        # Data rows
        row_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_values.append(row)

        # We'll just test some basic checks:
        # For the second row, 'col3' should have "something"
        # The actual column order may vary because we collect keys from a set.
        # We'll find indexes by headers:
        col1_index = headers.index("col1")
        col2_index = headers.index("col2") if "col2" in headers else None
        col3_index = headers.index("col3") if "col3" in headers else None

        # row_values[0] -> first data dict
        self.assertEqual(
            row_values[0][col1_index], "val1", "First row col1 should be 'val1'."
        )
        if col2_index is not None:
            self.assertEqual(
                row_values[0][col2_index], 10, "First row col2 should be 10."
            )

        # row_values[1] -> second data dict
        self.assertEqual(
            row_values[1][col1_index], "val2", "Second row col1 should be 'val2'."
        )
        if col3_index is not None:
            self.assertEqual(
                row_values[1][col3_index],
                "something",
                "Second row col3 should be 'something'.",
            )

    # -------------------------------------------------------------------------
    # _flatten_dict Tests
    # -------------------------------------------------------------------------

    def test_flatten_dict_deep_nesting_exceeds_max_depth(self):
        """Test that flattening raises an exception if max_depth is exceeded."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_deep"
        )
        # Prepare a nested dict deeper than default max_depth=10
        nested = {}
        current_level = nested
        for i in range(12):
            current_level["level"] = {}
            current_level = current_level["level"]

        with self.assertRaises(Exception) as cm:
            ds._flatten_dict(nested)

        self.assertIn("Maximum nesting depth", str(cm.exception))

    def test_flatten_dict_key_collision(self):
        """Test that a key collision raises an exception."""
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_collision"
        )
        # Suppose flattening results in the same key: 'a_b' appears in two ways
        data_with_collision = {"a": {"b": 1}, "a_b": 2}
        with self.assertRaises(Exception) as cm:
            ds._flatten_dict(data_with_collision)

        self.assertIn("Key collision detected", str(cm.exception))

    # -------------------------------------------------------------------------
    # Additional Edge Tests
    # -------------------------------------------------------------------------

    def test_folder_creation_failure(self):
        """
        Test that an OSError is raised if folder creation fails.
        We simulate this by passing an invalid directory name on Windows or
        a path with restricted permissions on other systems.
        """
        # We'll try an obviously invalid folder name (e.g., on most OS: /?<>|)
        # If this test fails on your OS, you might need a different approach or path.
        invalid_folder = os.path.join(self.test_dir, "invalid_folder_?:<>|")
        # Some OS might allow creation. We just do our best to test.
        # We'll catch OSError. If the OS allows it, that might lead to a test skip.
        try:
            with self.assertRaises(OSError):
                DataSaver(
                    folder=invalid_folder, file_extension="json", file_name="test"
                )
        except AssertionError:
            self.skipTest(
                "Folder creation did not fail on this OS with the given invalid path."
            )

    # -------------------------------------------------------------------------
    # __init__ coverage for multiple increments of i += 1
    # -------------------------------------------------------------------------
    def test_init_no_replace_file_multiple_increments(self):
        """
        Force DataSaver to increment the file name more than once.
        We create 'my_data.json' and 'my_data_1.json' ahead of time,
        so the saver should end up at 'my_data_2.json'.
        """
        base_path = os.path.join(self.test_dir, "my_data.json")
        base_path_1 = os.path.join(self.test_dir, "my_data_1.json")

        # Create two files
        with open(base_path, "w", encoding="utf-8") as f:
            f.write("File 0")
        with open(base_path_1, "w", encoding="utf-8") as f:
            f.write("File 1")

        ds = DataSaver(
            folder=self.test_dir,
            file_extension="json",
            replace_file=False,
            file_name="my_data",
        )
        self.assertTrue(
            ds.full_path.endswith("my_data_2.json"),
            "DataSaver should have incremented the name to my_data_2.json.",
        )

    # -------------------------------------------------------------------------
    # _save_json() exceptions coverage
    # -------------------------------------------------------------------------
    def test_save_json_type_error(self):
        """
        Provide a non-JSON-serializable object to trigger TypeError in _save_json.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="json", file_name="test_json_type_err"
        )
        data = [{"func": lambda x: x}]  # Lambdas are not JSON-serializable
        with self.assertRaises(TypeError) as cm:
            ds.save(data)
        self.assertIn("Failed to JSON-serialize the data", str(cm.exception))

    @patch("json.dump", side_effect=Exception("Unexpected error"))
    def test_save_json_unexpected_error(self, mock_json_dump):
        """
        Mock json.dump to raise a generic Exception, hitting the RuntimeError in _save_json.
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="json",
            file_name="test_json_unexpected_err",
        )
        with self.assertRaises(RuntimeError) as cm:
            ds.save([{"test": "data"}])
        self.assertIn("Unexpected error while saving JSON data", str(cm.exception))

    # -------------------------------------------------------------------------
    # _save_csv() exceptions coverage
    # -------------------------------------------------------------------------
    def test_save_csv_no_fieldnames(self):
        """
        Passing data with no keys triggers 'No valid field names found for CVS output.'
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="csv",
            file_name="test_csv_no_fieldnames",
        )
        data = [{}, {}]  # No keys => no field names
        with self.assertRaises(Exception) as cm:
            ds.save(data)
        self.assertIn("No valid field names found for CVS output.", str(cm.exception))

    @patch("csv.DictWriter.writerow", side_effect=Exception("Unexpected CSV error"))
    def test_save_csv_unexpected_error(self, mock_writerow):
        """
        Mock writer.writerow to raise an unexpected Exception, ensuring _save_csv hits RuntimeError.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="csv", file_name="test_csv_unexpected"
        )
        with self.assertRaises(RuntimeError) as cm:
            ds.save([{"col": "val"}])
        self.assertIn("Unexpected error while saving CSV data", str(cm.exception))

    @patch("csv.DictWriter.writerow", side_effect=csv.Error("Mocked CSV Error"))
    def test_save_csv_csv_error(self, mock_writerow):
        """
        Mock writer.writerow to raise a csv.Error, ensuring _save_csv hits csv.Error block.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="csv", file_name="test_csv_csv_err"
        )
        with self.assertRaises(csv.Error) as cm:
            ds.save([{"col": "val"}])
        self.assertIn("Mocked CSV Error", str(cm.exception))
        self.assertIn("CSV error while writing to", str(cm.exception))

    # -------------------------------------------------------------------------
    # _save_yaml() exceptions coverage
    # -------------------------------------------------------------------------
    @patch("yaml.safe_dump", side_effect=Exception("Unexpected YAML error"))
    def test_save_yaml_unexpected_error(self, mock_yaml_dump):
        """
        Mock yaml.safe_dump to raise a generic Exception, hitting the RuntimeError block in _save_yaml.
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="yaml",
            file_name="test_yaml_unexpected",
        )
        with self.assertRaises(RuntimeError) as cm:
            ds.save([{"key": "val"}])
        self.assertIn("Unexpected error while saving YAML data", str(cm.exception))

    @patch("yaml.safe_dump", side_effect=BaseException("Mocked YAMLError"))
    def test_save_yaml_yaml_error(self, mock_yaml_dump):
        """
        Mock yaml.safe_dump to raise a YAML-related error. Note: In practice, you'd raise `yaml.YAMLError`,
        but for mocking, we can simulate it.
        """
        # We can do: side_effect=yaml.YAMLError("Mocked YAMLError") to be more precise.
        ds = DataSaver(
            folder=self.test_dir, file_extension="yaml", file_name="test_yaml_yaml_err"
        )
        from yaml import YAMLError

        mock_yaml_dump.side_effect = YAMLError("Mocked YAMLError")

        with self.assertRaises(YAMLError) as cm:
            ds.save([{"key": "val"}])
        self.assertIn("YAML error while writing to", str(cm.exception))

    # -------------------------------------------------------------------------
    # _save_xml() exceptions coverage
    # -------------------------------------------------------------------------
    @patch(
        "xml.etree.ElementTree.ElementTree.write",
        side_effect=Exception("Unexpected XML Error"),
    )
    def test_save_xml_unexpected_error(self, mock_xml_write):
        """
        Mock ElementTree.write to raise a generic Exception, hitting the RuntimeError block in _save_xml.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="xml", file_name="test_xml_unexpected"
        )
        with self.assertRaises(RuntimeError) as cm:
            ds.save([{"key": "val"}])
        self.assertIn("Unexpected error while saving XML data", str(cm.exception))

    @patch("xml.etree.ElementTree.ElementTree.write")
    def test_save_xml_parse_error(self, mock_xml_write):
        """
        Mock ElementTree.write to raise ET.ParseError.
        """
        import xml.etree.ElementTree as ET

        mock_xml_write.side_effect = ET.ParseError("Mocked Parse Error")

        ds = DataSaver(
            folder=self.test_dir, file_extension="xml", file_name="test_xml_parse_err"
        )
        with self.assertRaises(ET.ParseError) as cm:
            ds.save([{"key": "val"}])
        self.assertIn("XML parsing error while writing to", str(cm.exception))

    # -------------------------------------------------------------------------
    # _save_pickle() exceptions coverage
    # -------------------------------------------------------------------------

    def test_save_pickle_attribute_error(self):
        """
        Pass an object that can't be pickled (like a local function or lambda) to raise AttributeError.
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="pkl",
            file_name="test_pickle_attribute_err",
        )
        data = [
            {"func": lambda x: x}
        ]  # Lambdas typically raise a Attribute during pickling
        with self.assertRaises(AttributeError) as cm:
            ds.save(data)
        self.assertIn("Cannot pickle data", str(cm.exception))

    @patch("pickle.dump", side_effect=Exception("Unexpected pickle error"))
    def test_save_pickle_unexpected_error(self, mock_pickle_dump):
        """
        Mock pickle.dump to raise a generic Exception, ensuring we hit the RuntimeError block.
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="pkl",
            file_name="test_pickle_unexpected_err",
        )
        with self.assertRaises(RuntimeError) as cm:
            ds.save([{"key": "val"}])
        self.assertIn("Unexpected error while saving Pickle data", str(cm.exception))

    # -------------------------------------------------------------------------
    # _save_xlsx() exceptions coverage
    # -------------------------------------------------------------------------
    def test_save_xlsx_no_valid_columns(self):
        """
        Pass data with no columns to trigger "No valid columns found for XLSX output.".
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="xlsx",
            file_name="test_xlsx_no_columns",
        )
        data = [{}, {}]  # No keys => no columns
        with self.assertRaises(Exception) as cm:
            ds.save(data)
        self.assertIn("No valid columns found for XLSX output.", str(cm.exception))

    @patch(
        "openpyxl.workbook.workbook.Workbook.save",
        side_effect=PermissionError("Mocked PermissionError"),
    )
    def test_save_xlsx_permission_error(self, mock_wb_save):
        """
        Mock Workbook.save to raise PermissionError.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="xlsx", file_name="test_xlsx_perm_err"
        )
        with self.assertRaises(PermissionError) as cm:
            ds.save([{"col": "val"}])
        self.assertIn(
            f"Permission denied while attempting to write the .xlsx file '{ds.full_path}'",
            str(cm.exception),
        )

    @patch(
        "openpyxl.workbook.workbook.Workbook.save",
        side_effect=OSError("Mocked OSError"),
    )
    def test_save_xlsx_os_error(self, mock_wb_save):
        """
        Mock Workbook.save to raise OSError.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="xlsx", file_name="test_xlsx_os_err"
        )
        with self.assertRaises(OSError) as cm:
            ds.save([{"col": "val"}])
        self.assertIn(
            "An OS error occurred while writing the .xlsx file", str(cm.exception)
        )

    @patch(
        "openpyxl.workbook.workbook.Workbook.save",
        side_effect=Exception("Unexpected XLSX error"),
    )
    def test_save_xlsx_unexpected_error(self, mock_wb_save):
        """
        Mock Workbook.save to raise a generic Exception, ensuring we hit the RuntimeError block.
        """
        ds = DataSaver(
            folder=self.test_dir,
            file_extension="xlsx",
            file_name="test_xlsx_unexpected_err",
        )
        with self.assertRaises(RuntimeError) as cm:
            ds.save([{"col": "val"}])
        self.assertIn("Unexpected error while saving XLSX data", str(cm.exception))

    @patch(
        "openpyxl.worksheet.worksheet.Worksheet.cell",
        side_effect=KeyError("Mocked Key Error"),
    )
    def test_save_xlsx_key_error(self, mock_cell):
        """
        Mock Worksheet.cell to raise KeyError while filling Excel row,
        ensuring we hit the 'Key error while filling Excel row' exception.
        """
        ds = DataSaver(
            folder=self.test_dir, file_extension="xlsx", file_name="test_xlsx_key_err"
        )
        with self.assertRaises(Exception) as cm:
            ds.save([{"col": "val"}])
        self.assertIn(
            "Key error while filling Excel row. Missing key:", str(cm.exception)
        )


# If you want to run these tests directly:
if __name__ == "__main__":
    unittest.main()
