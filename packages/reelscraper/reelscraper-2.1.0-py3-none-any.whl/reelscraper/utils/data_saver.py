from typing import List, Dict, Any
import xml.etree.ElementTree as ET
import os
import json
import csv
import yaml
import pickle
from openpyxl import Workbook


class DataSaver:
    def __init__(
        self,
        file_extension: str,
        replace_file: bool = False,
        folder: str = "data",
        file_name: str = "data",
    ) -> None:
        """
        Initialize DataSaver with folder location and file-related settings.

        :param folder: The folder path where files will be saved.
        :param file_extension: The format extension (e.g. 'json', 'csv', 'yml', 'yaml', 'xml', 'pkl', 'xlsx').
        :param replace_file: Whether to overwrite an existing file (default False).
        :param file_name: The base name of the file (without extension).
        :raises ValueError: If any of the arguments are invalid.
        :raises OSError: If the folder cannot be created or accessed.
        """
        if not folder or not isinstance(folder, str):
            raise ValueError(
                "A valid folder path must be provided as a non-emptry string."
            )

        if not file_extension or not isinstance(file_extension, str):
            raise ValueError(
                "A valid file extension must be provided as a non-empty string."
            )

        file_extension = file_extension.lstrip(".")

        if not file_extension:
            raise ValueError(
                "The file extension cannot be empty after removing leading periods."
            )

        file_name_without_ext, ext_in_name = os.path.splitext(file_name)

        if ext_in_name:
            raise ValueError(
                f"File name '{file_name}' already contains an extension '{ext_in_name}'. "
                "Please provide a base file name without any extension."
            )

        if not file_name or not isinstance(file_name, str):
            raise ValueError(
                "A valid file name must be provided as a non-empty string."
            )

        # Creates folder
        try:
            os.makedirs(folder, exist_ok=True)
        except OSError as e:
            raise OSError(f"Could not create or access folder '{folder}': {e}") from e

        self.full_path = os.path.join(folder, f"{file_name}.{file_extension.lower()}")

        if os.path.exists(self.full_path):
            if replace_file:
                pass
            else:
                # Create a new filename
                base_name = file_name
                i = 1
                while True:
                    new_file_name = f"{base_name}_{i}"
                    new_full_path = os.path.join(
                        folder, f"{new_file_name}.{file_extension.lower()}"
                    )
                    if not os.path.exists(new_full_path):
                        self.full_path = new_full_path
                        break
                    i += 1

        # Saves strategies implemented
        strategies = {
            "json": self._save_json,
            "csv": self._save_csv,
            "yml": self._save_yaml,
            "yaml": self._save_yaml,
            "xml": self._save_xml,
            "pkl": self._save_pickle,
            "xlsx": self._save_xlsx,
        }

        self.saver = strategies.get(file_extension.lower())

        if not self.saver:
            raise ValueError(
                f"Extension '{file_extension}' is not supported. "
                f"Supported formats: {', '.join(strategies.keys())}"
            )

    def save(self, data: List[Dict]) -> None:
        """
        Save the given data (list of dictionaries) to a file in the specified format.

        :param data: The list of dictionaries to save.
        :raises ValueError: If the file extension is not supported or data is invalid.
        :raises PermissionError: If permission is denied while writing the file.
        :raises FileNotFoundError: If the folder path is invalid.
        :raises OSError: If an OS error occurs while writing to the file.
        :raises RuntimeError: If an unexpected error occurs.
        """
        # Checks for data correct input
        if not data or not isinstance(data, list):
            raise ValueError(
                "Data must be provided as a non-empty list of dictionaries."
            )
        for i, item in enumerate(data):
            if not isinstance(item, dict):
                raise ValueError(
                    f"Item at intex {i} is not a dictionary. All items must be dictionaries."
                )

        self.saver(data)

    def _save_json(self, data: List[Dict]) -> None:
        """
        Saves data in .json format
        """
        try:
            with open(self.full_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except TypeError as e:
            # If data contains non-serializable objects
            raise TypeError(f"Failed to JSON-serialize the data: {e}") from e
        except Exception as e:
            raise RuntimeError(f"Unexpected error while saving JSON data: {e}") from e

    def _save_csv(self, data: List[Dict]) -> None:
        """
        Saves data in .csv format
        """
        fieldnames = self._get_csv_fieldnames(data)
        if not fieldnames:
            raise ValueError("No valid field names found for CVS output.")

        try:
            with open(self.full_path, "w", newline="", encoding="utf-8") as f:
                writer: csv.DictWriter = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for item in data:
                    writer.writerow(self._flatten_dict(item))
        except csv.Error as e:
            raise csv.Error(
                f"CSV error while writing to '{self.full_path}': {e}"
            ) from e
        except Exception as e:
            raise RuntimeError(f"Unexpected error while saving CSV data: {e}") from e

    def _save_yaml(self, data: List[Dict]) -> None:
        """
        Saves data in .yaml or .yml format
        """
        try:
            with open(self.full_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(data, f, sort_keys=False, allow_unicode=True)
        except yaml.YAMLError as e:
            raise yaml.YAMLError(
                f"YAML error while writing to '{self.full_path}': {e}"
            ) from e
        except Exception as e:
            raise RuntimeError(f"Unexpected error while saving YAML data: {e}") from e

    def _save_xml(self, data: List[Dict]) -> None:
        """
        Saves data in .xml, with tag <root> and child tag <record>
        """
        root: ET.Element = ET.Element("root")

        for item in data:
            record_elem: ET.SubElement = ET.SubElement(root, "record")
            for key, value in self._flatten_dict(item).items():
                field_elem: ET.SubElement = ET.SubElement(record_elem, key)
                field_elem.text = str(value) if value is not None else ""

        tree: ET.ElementTree = ET.ElementTree(root)
        try:
            tree.write(self.full_path, encoding="utf-8", xml_declaration=True)
        except ET.ParseError as e:
            raise ET.ParseError(
                f"XML parsing error while writing to '{self.full_path}': {e}"
            ) from e
        except Exception as e:
            raise RuntimeError(f"Unexpected error while saving XML data: {e}") from e

    def _save_pickle(self, data: List[Dict]) -> None:
        """
        Saves data in .pkl format
        """
        try:
            with open(self.full_path, "wb") as f:
                pickle.dump(data, f)
        except AttributeError as e:
            # This can occur if the data contains objects that can't be pickled
            raise AttributeError(f"Cannot pickle data: {e}") from e
        except Exception as e:
            raise RuntimeError(f"Unexpected error while saving Pickle data: {e}") from e

    def _save_xlsx(self, data: List[Dict]) -> None:
        """
        Saves data in .xlsx format
        """
        flattened_data = [self._flatten_dict(item) for item in data]
        columns = list({col for item in flattened_data for col in item.keys()})
        if not columns:
            raise ValueError("No valid columns found for XLSX output.")

        wb: Workbook = Workbook()
        ws = wb.active
        ws.title = "Data"

        try:
            # Write header row
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)

                # Write data rows
                for row_idx, item in enumerate(flattened_data, start=2):
                    for col_idx, col_name in enumerate(columns, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(col_name))
        except KeyError as e:
            raise KeyError(
                f"Key error while filling Excel row. Missing key: {e}"
            ) from e

        try:
            wb.save(self.full_path)
        except PermissionError as e:
            raise PermissionError(
                f"Permission denied while attempting to write the .xlsx file '{self.full_path}'."
            ) from e
        except OSError as e:
            raise OSError(
                f"An OS error occurred while writing the .xlsx file '{self.full_path}': {e}"
            ) from e
        except Exception as e:
            raise RuntimeError(f"Unexpected error while saving XLSX data: {e}") from e

    def _get_csv_fieldnames(self, data_list: List[Dict]) -> List[str]:
        """
        Collects all CSV fields from the data
        """
        field_set = set()
        for item in data_list:
            field_set.update(self._flatten_dict(item).keys())

        return list(field_set)

    def _flatten_dict(
        self,
        nested_data_dict: Dict[str, Any],
        parent_key: str = "",
        separator: str = "_",
        max_depth: int = 10,
        current_depth: int = 0,
    ) -> Dict[str, Any]:
        """
        Flattens a nested dictionary into a single-level dictionary.
        :param nested_data_dict: Dizionario annidato da appiattire.
        :param parent_key: Prefisso di chiave.
        :param separator: Separatore tra i livelli di chiave.
        :param max_depth: Massima profondità di ricorsione consentita.
        :param current_depth: Profondità attuale di ricorsione (interna).
        :raises FlattenDictError: Se si supera la profondità massima o si verifica una collisione di chiavi.
        """
        if current_depth > max_depth:
            raise Exception(
                f"Maximum nesting depth ({max_depth}) exceeded when flattening dictionary."
            )

        items = {}
        for key, value in nested_data_dict.items():
            new_key = f"{parent_key}{separator}{key}" if parent_key else key
            if isinstance(value, dict):
                try:
                    sub_dict = self._flatten_dict(
                        value,
                        parent_key=new_key,
                        separator=separator,
                        max_depth=max_depth,
                        current_depth=current_depth + 1,
                    )
                except Exception:
                    raise

                for k_sub, v_sub in sub_dict.items():
                    if k_sub in items:
                        raise Exception(
                            f"Key collision detected: '{k_sub}' already exists."
                        )
                    items[k_sub] = v_sub
            else:
                if new_key in items:
                    raise Exception(
                        f"Key collision detected: '{new_key}' already exists."
                    )
                items[new_key] = value

        return items
