#!/usr/bin/python
# -*- coding: UTF-8 -*-
# @Time    : 2021/10/21 9:32 下午
# @Author  : xinlan
import os
import sys
from argparse import ArgumentParser
from configparser import ConfigParser
import yaml
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from .runner import TestRunner
from .common.log_handler import get_logger
from .common.db_handler import DB
from .case import BaseCase
from .loader import TestLoader
from .utils import DjangoJSONEncoder


class TestProgram:
    def __init__(self, args=None, test_data=None, url=None):
        self.test_data = test_data
        self.result = None
        self.args = None
        self.parse_args(args)
        self.test_base_dir = None
        self.settings = {}
        self.get_settings()
        self.logger = get_logger(file=self.settings.get('logfile'), debug=self.settings.get('debug', False))
        if self.test_data is None:
            self.test_data = self.get_test_data_from_file()
        else:
            self.__get_test_data()
        self.TestLoader = TestLoader
        self.TestRunner = TestRunner
        if self.settings.get('auto') == 1:
            self.run_tests()
            print('用例总数:{},成功:{}个,跳过:{},失败:{}个,错误:{}个'.format(
                self.result['total'],
                self.result['passed'],
                self.result['skipped'],
                self.result['failed'],
                self.result['broken']
            ))
            exit(self.result['status'])

    def __get_test_data(self):
        if isinstance(self.test_data, dict):
            test_suits = self.test_data.pop('test_suits')
            self.settings.update(self.test_data)
            self.test_data = test_suits

    def compile_result(self):
        result = {
            'total': 0,
            'status': 0,
            'failed': 0,
            'broken': 0,
            'passed': 0,
            'skipped': 0,
            'test_data': [],
            'duration': 0

        }
        for item in self.result:
            temp = item.result()
            result['total'] += temp.testsRun
            result['failed'] += len(temp.failures)
            result['broken'] += len(temp.errors)
            result['passed'] += temp.passed
            result['skipped'] += len(temp.skipped)
            result['duration'] += temp.duration
            result['test_data'].append(
                {'name': temp.name,
                 'result': temp.result
                 }
            )

        if result['broken'] > 0:
            result['status'] = 2
        elif result['failed'] > 0:
            result['status'] = 1
        result['total'] = result['broken'] + result['failed'] + result['passed'] + result['skipped']
        self.result = result

    def run_tests(self):
        self.result = []
        with ThreadPoolExecutor(max_workers=self.settings.get('thread_num', 1)) as executor:
            for ts in self.create_test():
                self.result.append(executor.submit(self.TestRunner().run, ts))
        self.compile_result()

    def create_test(self):
        for item in self.test_data:
            if not item.get('testcases'):
                continue
            self.add_info_to_data(item)
            yield self.TestLoader(data=item, baseclass=BaseCase).loader()

    def add_info_to_data(self, data):
        data['settings'] = self.settings
        data['logger'] = self.logger.getChild(data['name'])
        data['db'] = DB(**self.settings.get('db_config')) if self.settings.get('db_config', None) else None

    def parse_args(self, args):
        if sys.version < '3.9':
            parser = ArgumentParser(description='easytest 命令行参数帮助')
        else:
            parser = ArgumentParser(description='easytest 命令行参数帮助', exit_on_error=False)
        parser.add_argument('file_or_dir', nargs='?', type=str, help='项目路径，或者需要执行的用例文件')
        parser.add_argument('--debug', action="store_true", help='开启日志调试模式，默认为False')
        parser.add_argument('--logfile', type=str, help='日志文件路径')
        parser.add_argument('--marks', type=str, default='', help='运行标记')
        parser.add_argument('--thread_num', type=int, default=0, help='运行启动线程的数量')
        parser.add_argument('--report', type=str, help='测试报告文件路径，按文件后缀生成对应格式的报告')
        parser.add_argument('--auto', type=int, default=1, help='是否自动执行用例，还是要主动运行方法运行，默认为1表示自动运行')
        if args:
            self.args = vars(parser.parse_args(args))
        else:
            self.args = vars(parser.parse_args())

    def get_settings(self):
        ini_file = os.path.join(os.getcwd(), 'easytest.ini')
        if os.path.exists(ini_file):
            self.settings = self.get_ini_config(ini_file)

        for key, value in self.args.items():
            if value:
                self.settings[key] = value
        self.settings['file_or_dir'] = self.args['file_or_dir']

    @staticmethod
    def get_ini_config(ini_file):
        config = ConfigParser()
        try:
            config.read(ini_file)
        except UnicodeDecodeError:
            try:
                config.read(ini_file, encoding='utf-8')
            except UnicodeDecodeError as e:
                raise e
        res = {
        }
        for section in config.sections():
            if section in ['project', 'run']:
                res.update(dict(config[section]))
            else:
                res[section] = dict(config[section])
        if 'run' in config:
            if 'debug' in config['run']:
                res['debug'] = config['run'].getboolean('debug')
            if 'thread_num' in config['run']:
                res['thread_num'] = config['run'].getint('thread_num')
            if 'retry' in config['run']:
                res['retry'] = config['run'].getint('retry')
        if 'db_config' in config and 'port' in config['db_config']:
            res['db_config']['port'] = config['db_config'].getint('port')
        return res

    def get_test_data_file_name(self):
        files = []
        file = self.args['file_or_dir']
        if file is None:
            file = os.getcwd()
            self.test_base_dir = file
        file = os.path.abspath(file)

        if os.path.isfile(file):
            files.append(file)
            self.test_base_dir = os.path.dirname(file)
        elif os.path.isdir(file):
            self.test_base_dir = file
            for root, _, fs in os.walk(file):
                for f in fs:
                    if os.path.basename(f).startswith('~'):
                        continue
                    if f.split('.')[-1] in ['xlsx', 'yaml', 'yml']:
                        files.append(os.path.join(root, f))
        else:
            self.logger.warning('目录/文件：{}不存在'.format(file))
            raise RuntimeError('目录/文件：{}不存在'.format(file))

        return files

    def get_test_data_from_file(self):
        files = self.get_test_data_file_name()

        for file in files:
            if file.split('.')[-1] == 'xlsx':
                for data in self.load_data_from_excel(file):
                    yield data
            else:
                for data in self.load_data_from_yaml(file):
                    yield data

    def load_data_from_excel(self, file):
        """
        从excel文件中获取测试数据
        :param file:
        :return:
        """
        wb = load_workbook(file)
        sheet_names = wb.sheetnames
        name_prefix = self.get_name_prefix(file)

        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            suit = {
                "name": ':'.join((name_prefix + [sheet_name])),
                'path': file,
                "testcases": []
            }
            row = sheet.max_row
            column = sheet.max_column
            title = {}
            for i in range(1, column + 1):
                title[i] = sheet.cell(row=1, column=i).value

            for j in range(2, row + 1):
                temp = {}
                for i in range(1, column + 1):
                    temp[title[i]] = sheet.cell(row=j, column=i).value
                suit['testcases'].append(temp)
            yield suit

    def load_data_from_yaml(self, file):
        with open(file, 'rb') as f:
            temp = yaml.load(f, yaml.FullLoader)
        name_prefix = self.get_name_prefix(file)

        data = {'name': ':'.join(name_prefix)}
        if 'test' in temp:
            data.update({
                'path': file,
                'testcases': [temp['test']]
            })

        if 'test_suit' in temp:
            data.update({
                'path': file,
                'testcases': temp['test_suit']
            })

        yield data

    def get_name_prefix(self, filename):
        filename = ''.join(filename.split('.')[:-1])
        res = filename.replace(self.test_base_dir, '').strip('\\/')
        if '/' in res:
            res = res.split('/')
        else:
            res = res.split('\\')
        return res

    def generate_report(self):
        if self.settings['report'].split('.')[-1] == 'json':
            with open(self.settings['report'], 'w', encoding='utf-8') as f:
                import json
                json.dump(self.result, f, ensure_ascii=False, cls=DjangoJSONEncoder)
        else:
            print('正在开发中，敬请期待')

    def run(self):
        self.run_tests()

        return self.return_result()

    def return_result(self):
        if self.result is None:
            self.run_tests()
        return {
            'total_test_suit': len(self.result['test_data']),
            'total_test_case': self.result['total'],
            'test_suit_result': self.result['test_data'],
            'fail_test_case': self.result['failed'],
            'success_test_case': self.result['passed'],
            'error_test_case': self.result['broken'],
            'skip_test_case': self.result['skipped'],
            'duration': self.result['duration'],
            'status': self.result['status']
        }


main = TestProgram
