#! /usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2020/11/13 21:49
# @Author : 心蓝
import os
import re
import random
import json

import yaml
from openpyxl import load_workbook

# from common.db_handler import db


def get_data_from_excel(file, sheet_name=None):
    """
    从excel文件中获取测试用例数据
    :param file: excel文件名
    :param sheet_name: 表名
    :return:
    """
    # 1. 加载excel文件
    wb = load_workbook(file)
    # 2. 获取对应的表
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    # 获取几行几列
    row = ws.max_row
    column = ws.max_column
    # 3. 获取cell中的数据并组织成需要的结构
    # 创建一个空列表用来接收数据
    data = []
    # 先获取标题
    title = {}
    for i in range(1, column+1):
        title[i] = ws.cell(row=1, column=i).value
    # title = {i: ws.cell(row=1, column=i).value for i in range(1, column+1)}
        # print(ws.cell(row=1, column=i).value)
        # print(1, i)
    # print(title)
    # title = {
    #     1: 'title',
    #     2: 'username',
    #     3: 'password',
    #     4: 'expect'
    # }
    # 获取值
    # print(row, column)
    for j in range(2, row+1):
        # 临时变量，表示一行数据
        temp = {}
        for i in range(1, column+1):
            # print(j, i)

            # print(title[i], ws.cell(row=j, column=i).value, end=' ')
            temp[title[i]] = ws.cell(row=j, column=i).value
        # 循环结束的时候i=column 也就是最后一列 title[i]取出expect
        # print(temp)
        # print(temp[title[i]])
        # temp[title[i]] = eval(temp[title[i]])
        # print(temp)
        data.append(temp)

    return data


def load_data_from_excel(file):
    """
    从excel文件中获取测试数据
    :param file:
    :return:
    """
    wb = load_workbook(file)
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        suit = {
            "name": sheet_name,
            'path': file,
            "testcases": []
        }
        row = sheet.max_row
        column = sheet.max_column
        title = {}
        for i in range(1, column + 1):
            title[i] = sheet.cell(row=1, column=i).value

        for j in range(2, row + 1):
            temp = {}
            for i in range(1, column + 1):
                temp[title[i]] = sheet.cell(row=j, column=i).value
            suit['testcases'].append(temp)
        yield suit


def load_data_from_yaml(file):
    with open(file, 'rb') as f:
        temp = yaml.load(f, yaml.FullLoader)
    data = {'name': os.path.split(file)[-1].split('.')[0]}
    if 'test' in temp:
        data.update({
                'path': file,
                'testcases': [temp['test']]
              })

    if 'test_suit' in temp:
        data.update({
                'path': file,
                'testcases': temp['test_suit']
                })

    yield data


def generate_phone_num():
    """
    随机生成手机号码
    :return:
    """
    # 1开头，11位，第二数3456789，后面9位
    # 1开头，11位，第二个数358
    # 写死，避免后面出问题，如果要测手机号码单独测试
    phone = ['158']
    # 剩下的9位
    for i in range(8):
        phone.append(random.choice('0123456789'))
    phone = ''.join(phone)

    return phone


# def phone_is_exist(phone_num):
#     """
#     手机号码是否已注册
#     :param phone_num:
#     :return:
#     """
#     sql = "select id from member where mobile_phone='{}'".format(phone_num)
#     print(sql)
#     if db.exist(sql):
#         return True
#     return False


def replace_args(s, **kwargs):
    """
    替换参数
    :param s: 需要被替换的用例数据
    :param kwargs: 需要替换的参数
    :return: 替换好的用例数据
    """
    for key, value in kwargs.items():
        s = s.replace('#{}#'.format(key), str(value))
    return s


def replace_args_by_re(s, cls):
    """
    通过正则表达式动态替换参数
    :param s: 需要被替换的json字符串
    :param cls: 提供数据的对象
    :return:
    """
    args = re.findall('#(.*?)#', s)
    for arg in args:
        value = getattr(cls, arg, None)
        if value:
            s = s.replace('#{}#'.format(arg), str(value))
    return s


def extract_data(maps, cls, json_res_str):
    """
    提取需要存储的数据
    :param maps: 需要提取的映射 {cls_arg: key}
    :param cls: 测试用例类
    :param json_res_str: json响应字符串
    :return:
    """
    # 去掉干扰字符
    json_res_str = json_res_str.replace(' ', '').replace('\n', '')

    for arg, key in maps.items():
        # 根据key生成正则表达式
        re_strs = [r'"{}":"(.*?)"'.format(key),  # 字符串参数
                   r'"{}":(\d+?\.\d+?),'.format(key),  # 浮点数参数,放在前面
                   r'"{}":(\d+?),'.format(key),  # 数字参数
                   ]
        # 取响应中提取数据，并保存到类中
        for re_str in re_strs:
            value = re.findall(re_str, json_res_str)
            if value:
                setattr(cls, arg, value[0])
                # 找到就跳出
                break


if __name__ == '__main__':
    res = load_data_from_excel('../project.xlsx')
    # res = get_data_from_excel('../testdata/testcases.xlsx', sheet_name='register')
    # print(res[0]['request_data'].replace('#phone#', generate_phone_num()))
    # phone = generate_phone_num()
    # print(phone)
    # phone = generate_phone_num()
    # print(phone)
    # print(res[0]['expect'], type(res[0]['expect']))
    # print(random.random())  # 返回0-1之间的随机浮点数，左闭右开
    # print(random.randint(a, b)) # 返回整数a到b之间的随机整数，闭区间
    # random.choice()
    # 随机获取ls中的一个字母
    # ls = list('abcdefghijk')
    # 随机获取索引
    # ls的索引区间 0-len(ls)-1
    # random_index = random.randint(0, len(ls)-1)
    # print(ls[random.randint(0, len(ls)-1)])
    # print(random.choice(ls))

