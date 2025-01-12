import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Union

from jinja2 import Environment, PackageLoader

from django_scaffolding_tools.builders import build_serializer_data
from django_scaffolding_tools.exceptions import DSTException
from django_scaffolding_tools.parsers import parse_dict, parse_for_patterns, transform_dict_to_model_list
from django_scaffolding_tools.patterns import PATTERN_FUNCTIONS


class ReportWriter:
    def __init__(self):
        self.template_env = Environment(loader=PackageLoader("django_scaffolding_tools", "templates"))

    def write(self, template_name: str, output_file: Path, **params):
        template = self.template_env.get_template(template_name)
        output = template.render(**params)
        with open(output_file, "w") as html_file:
            html_file.write(output)


def write_serializer_from_file(source_file: Path, output_file: Path):
    with open(source_file) as json_file:
        data = json.load(json_file)

    writer = ReportWriter()
    parsed_dict = parse_dict(data)

    model_list = transform_dict_to_model_list(parsed_dict)
    model_list = parse_for_patterns(model_list, PATTERN_FUNCTIONS)
    model_list = build_serializer_data(model_list)

    writer.write("serializers.py.j2", output_file, model_list=model_list)


def get_keyword(att_keywords: List[Dict[str, Any]], keyword_name: str) -> Union[str, int]:
    for att_keyword in att_keywords:
        try:
            if att_keyword["name"] == keyword_name:
                return att_keyword["value"]
        except TypeError as e:
            error_message = f"Error getting keyword {keyword_name}. Error: {e}"
            raise DSTException(error_message)


def write_django_model_csv(models_list: List[Dict[str, Any]], filename: Path):
    with open(filename, "w") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["Class name", "Attribute", "Field type", "Max length", "Description"])
        for model in models_list:
            class_name = model["name"]
            for att in model["attributes"]:
                att_name = att["name"]
                att_field_type = att["data_type"]
                max_length = get_keyword(att["keywords"], "max_length")
                description = get_keyword(att["keywords"], "help_text")
                if description is not None:
                    description = description.replace(",", " ")
                writer.writerow([class_name, att_name, att_field_type, max_length, description])


def write_django_model_excel(models_list: List[Dict[str, Any]], filename: Path):
    headers = {
        "class_name": {"title": "Class name"},
        "attribute": {"title": "Attribute name"},
        "field_type": {"title": "Field type"},
        "max_value": {"title": "Max value"},
        "description": {"title": "Description"},
    }
    excel_data = list()
    for model in models_list:
        for att in model["attributes"]:
            excel_dict = {"class_name": model["name"], "attribute": att["name"], "field_type": att["data_type"]}
            max_value = get_keyword(att["keywords"], "max_length")
            if max_value is not None:
                max_value = f"max length={max_value}"
            else:
                max_digits = get_keyword(att["keywords"], "max_digits")
                decimal_places = get_keyword(att["keywords"], "decimal_places")
                if max_digits is not None:
                    max_value = f"digits={max_digits}, decimal places={decimal_places}"
            excel_dict["max_value"] = max_value
            excel_dict["description"] = get_keyword(att["keywords"], "help_text")
            excel_data.append(excel_dict)
    from django_scaffolding_tools.utils.core import quick_write

    quick_write(excel_data, filename.parent / "__excel.json")
    simple_write_to_excel(filename, headers, excel_data)


def simple_write_to_excel(filename: Path, headers: Dict[str, Any], lines: List[Dict[str, Any]]):
    """Writes to excel a List of dictionaries. The headers keys must match the keys of the values to write.
    The headers must contain a dictionary with the title key.
    """
    from openpyxl.workbook import Workbook

    wb = Workbook()
    sheet = wb.create_sheet()
    row = 1
    col = 1
    for key, header in headers.items():
        sheet.cell(row=row, column=col, value=header["title"])
        col += 1
    row += 1
    for line in lines:
        col = 1
        for key in headers:
            value = line.get(key)
            if value is not None:
                sheet.cell(row=row, column=col, value=value)
            col += 1
        row += 1

    wb.save(filename)


def write_enums(enum_name: str, enumerations: List[Dict[str, Any]], module_file: Path) -> None:
    lines = list()
    line = f"class {enum_name}(str, Enum):\n"
    lines.append(line)
    for enumeration in enumerations:
        line = f'\t{enumeration["name"]} = \'{enumeration["value"]}\'\n'
        lines.append(line)
    with open(module_file, "w") as m_file:
        m_file.writelines(lines)
