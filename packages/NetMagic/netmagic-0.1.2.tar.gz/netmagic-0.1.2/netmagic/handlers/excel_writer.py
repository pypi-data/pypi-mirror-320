"""
Netmagic Excel Writer
by Michael Buckley

This is for publishing the data collected from other modules to a human-readable format
using Microsoft Excel (which OpenOffice (freeware) can also read)

"""

from re import search

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

class CellEntry:
    def __init__(self, value, font: Font = None, fill: PatternFill = None, 
                 shape = None) -> None:
        self.value = value
        self.font = font
        self.fill = fill
        self.shape = shape


class RowEntry:
    def __init__(self, entries: list[CellEntry]) -> None:
        self.entries = entries

    def __repr__(self) -> str:
        return f'RowEntry (Len: {len(self.entries)})'

    def __iter__(self):
        for entry in self.entries:
            yield entry


class Section:
    def __init__(self, rows: list[RowEntry], style = None) -> None:
        self.rows = rows
        self.style = style

    def __repr__(self) -> str:
        return f'Section (Len: {len(self.rows)})'

    def apply_row_font(self, font: Font, row_number: int):
        for cell in self.rows[row_number]:
            cell.font = font


class SheetEntry:
    def __init__(self, name: str, sections: list[Section]) -> None:
        self.name = name
        self.sections = sections

    def __repr__(self) -> str:
        return f'SheetEntry (Len: {len(self.sections)})'


def handle_cell(cell: Cell, cell_entry: CellEntry):
    """
    Set value and apply formatting of a cell with `CellEntry` instance
    """
    # Aux to cover strings not enrolled in cel objects
    if isinstance(cell_entry, str):
        cell_entry = CellEntry(cell_entry)
        
    cell.value = str(cell_entry.value)
    
    if cell_entry.font:
        cell.font = cell_entry.font
    if cell_entry.fill:
        cell.fill = cell_entry.fill
        
def prepare_sheet(sheet: Worksheet, sections: list[Section]):
    """
    This collects info a type and prepares a sheet to be added to a workbork
    """
    x = 0

    for section in sections:
        for row in section.rows:
            x += 1
            y = 1
            for cell_entry in row:
                handle_cell(sheet.cell(x, y), cell_entry)
                y += 1
                
def prepare_workbook(filename: str, sheet_entries: list[SheetEntry]):
    """
    Prepare and save an Excel file from a series of pre-defined entries
    """
    if not search(r'\.xlsx$', filename):
        filename = f'{filename}.xlsx'

    workbook= Workbook()
    default_sheet = workbook.active

    for sheet_entry in sheet_entries:
        sheet= workbook.create_sheet(sheet_entry.name)
        prepare_sheet(sheet, sheet_entry.sections)

    if default_sheet.title in workbook.sheetnames:
        workbook.remove(default_sheet)

    workbook.save(filename)
