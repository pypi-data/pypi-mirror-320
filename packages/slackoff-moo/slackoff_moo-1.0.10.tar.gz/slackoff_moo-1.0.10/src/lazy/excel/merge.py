from os.path import isfile, isdir, join
from os import listdir
from xlrd import open_workbook
from openpyxl import load_workbook, Workbook
from ..toj.utils import Utils
import shutil


class XlsIterator:
    def __init__(self, workbook, params):
        self.xls_sheet = None
        self.workbook = workbook
        self.start_row_index = -1
        self.sheet_index = 0
        self.sheet_names = []
        self.sheet_current_row = -1  # 当前读取表格的行数
        self.sheet_rows = -1  # 表格行数

        if "sheet" not in params:
            raise Exception("No sheet name")
        self.params = params
        self.sheet_params = params["sheet"]
        if "start_row" in self.sheet_params:
            self.start_row_index = self.sheet_params["start_row"]
        if "filter" not in self.sheet_params:
            self.sheet_params["filter"] = lambda s: True
        if "sheet_name" in self.sheet_params:
            self.sheet_names.append(self.sheet_params["sheet_name"])
        else:
            for sheet_name in workbook.sheet_names():
                if Utils.handle(
                        self.sheet_params["sheet_name_filter"], "sheet_name_filter",
                        {"sheet_name": sheet_name, "filename": params["filename"]}
                ):
                    self.sheet_names.append(sheet_name)

    def __iter__(self):
        return self

    def __next__(self):
        while True:
            row = self.next()
            if Utils.handle(
                    self.sheet_params["sheet_row_filter"], "sheet_row_filter",
                    {"row": row, "filename": self.params["filename"], "sheet_name": self.sheet_names[self.sheet_index - 1]}
            ):
                return row

    def next(self, ):
        self.sheet_current_row += 1
        if self.sheet_current_row >= self.sheet_rows:
            if self.sheet_index >= len(self.sheet_names):
                raise StopIteration
            self.xls_sheet = self.workbook.sheet_by_name(self.sheet_names[self.sheet_index])
            self.sheet_rows = self.xls_sheet.nrows
            self.sheet_current_row = self.start_row_index - 1
            print(f"读取 \r{self.params['filename'] + ':' + self.sheet_names[self.sheet_index]} ")
            self.sheet_index += 1
        row = [self.xls_sheet.cell_value(self.sheet_current_row, cols) for cols in range(self.xls_sheet.ncols)]
        return row

    def __del__(self):
        self.workbook.release_resources()


class XlsxIterator:
    def __init__(self, workbook, params):
        self.workbook = workbook
        if "sheet" not in params:
            raise Exception("No sheet name")
        sheet_params = params["sheet"]
        if "sheet_name" in sheet_params:
            self.sheet = workbook.sheet_by_name(sheet_params["sheet_name"])
        else:
            self.sheet = workbook.sheet_by_index(0)
        if "start_row" in sheet_params:
            self.iterator = self.sheet.iter_rows(min_row=sheet_params["start_row"])
        else:
            self.iterator = self.sheet.iter_rows(min_row=1)

    def __iter__(self):
        return self

    def __next__(self):
        return [item.value for item in self.iterator]

    def __del__(self):
        self.workbook.close()


class ExcelMergeTarge:
    def __init__(self, params):
        if "target" not in params:
            raise ValueError()
        target = params["target"]
        if "path" not in target:
            if "folder" not in target and "filename" not in target:
                raise ValueError()
            target["path"] = join(params["folder"], params["filename"])
        self.path = target["path"]
        if "template" in target:
            self.copy(target["template"], target["path"])
        if target["path"].endswith('.xlsx'):
            if "sheet_name" not in target:
                self.workbook = load_workbook(filename=target["path"])
                self.sheet = self.workbook.active
            else:
                self.workbook = load_workbook(filename=target["path"])
                self.sheet = self.workbook[target["sheet_name"]]
        else:
            raise ValueError("目标文件只能是xlsx格式的表格")

    def save(self, row):
        self.sheet.append(row)

    def __del__(self):
        self.workbook.save(self.path)
        self.workbook.close()

    @staticmethod
    def copy(source, target):
        if source.endswith('.xls'):
            source_wk = open_workbook(filename=source, on_demand=True)
            source_ws = source_wk.sheet_by_index(0)
            wb = Workbook()
            ws = wb.active
            for row in range(source_ws.nrows):
                values = [source_ws.cell_value(row, col) for col in range(source_ws.ncols)]
                ws.append(values)
            wb.save(target)
            wb.close()
            source_wk.release_resources()
        else:
            shutil.copy(source, target)


class ExcelMerge:
    @staticmethod
    def valid(params):
        pass

    @staticmethod
    def xls_iterator(frame):
        filepath = frame["filename"]
        source_work = open_workbook(filename=filepath, on_demand=True)
        return XlsIterator(source_work, frame)

    @staticmethod
    def xlsx_interator(frame):
        filepath = frame["filename"]
        source_workbook = load_workbook(filename=filepath)
        return XlsxIterator(source_workbook, frame)

    def excel_iterator(self, frame):
        filepath = frame["filename"]
        if filepath.endswith('.xls'):
            return self.xls_iterator(frame)
        elif filepath.endswith('.xlsx'):
            return self.xls_iterator(frame)
        return None

    def merge_excel(self, params):
        self.valid(params)
        sources = params["sources"]
        target = ExcelMergeTarge(params)
        for source in sources:
            stack = [item for item in source["path"]]
            sheet_frame = []
            while len(stack) > 0:
                path = stack.pop()
                if isfile(path):
                    temp_source = source.copy()
                    temp_source["filename"] = path
                    sheet_frame.append(temp_source)
                if isdir(path):
                    for item in listdir(path):
                        stack.append(join(path, item))

            for frame in sheet_frame:
                iterator = self.excel_iterator(frame)
                if iterator is not None:
                    if isinstance(iterator, object) and hasattr(iterator, "__iter__") and hasattr(iterator, "__next__"):
                        for row in iterator:
                            target.save(row)
