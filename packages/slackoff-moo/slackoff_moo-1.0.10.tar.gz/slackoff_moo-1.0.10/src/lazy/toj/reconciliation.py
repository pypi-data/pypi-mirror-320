from abc import ABC, abstractmethod
from .reconciliation_source import XlsxDataSource
from .reconciliation_rule import OrganizationRule
from .reconcilization_merge import DefaultMerge
from .reconciliation_to_excel import ToExcel
from openpyxl import load_workbook


# 配置整合入口
class Reconciliation(ABC):
    @abstractmethod
    def handle(self):
        pass


# 默认生成数据源的迭代器
class SourceIterator:
    def __init__(self, min_row, sheet_name):
        self.min_row = min_row
        self.sheet_name = sheet_name

    def init_iterator(self, ctx):
        if ctx.current_file_index >= len(ctx.files):
            ctx.iter = None
        else:
            filepath = ctx.files[ctx.current_file_index]
            ctx.current_file_index += 1
            workbook = load_workbook(filename=filepath)

            if self.sheet_name is None:
                sheet = workbook.active
                ctx.iter = sheet.iter_rows(min_row=self.min_row)
            else:
                sheet = workbook[self.sheet_name]
                ctx.iter = sheet.iter_rows(min_row=self.min_row)


# 默认实现，读取的数据源必须是openpyxl
class XlsxReconciliation(Reconciliation):
    def __init__(self, rules_):
        if "generator" not in rules_:
            raise ValueError("rules must contain 'generator'")
        if isinstance(rules_["generator"], dict):
            rules_["generator"] = [rules_["generator"]]
        self.rules = rules_

    def handle(self):
        load_sources = []
        for g in self.rules["generator"]:
            source = g["source"]

            # 适配openpyxl的默认配置
            if "iterator_item" not in source:
                source["iterator_item"] = lambda cells: [cell.value for cell in cells]
            if "init_iterator" not in source:
                min_row = 1
                sheet_name = None
                if "sheet" in source:
                    if "min_row" in source["sheet"]:
                        min_row = source["sheet"]["min_row"]
                    if "sheet_name" in source["sheet"]:
                        sheet_name = source["sheet"]["sheet_name"]
                source["init_iterator"] = SourceIterator(min_row=min_row, sheet_name=sheet_name)

            iterator = XlsxDataSource(g["source"])
            organization = OrganizationRule(g["rules"], iterator)  # 迭代器
            load_sources.append(organization.parse({"row": iterator}))  # iterator是一个迭代器
        if "merge" in self.rules:
            # 合并数据
            if len(load_sources) == 0:
                raise ValueError("no sources found")
            while len(load_sources) == 1:
                load_sources.append({})
            defaultMerge = DefaultMerge(self.rules["merge"])
            load_sources.append(defaultMerge.merges(load_sources))
        if "toXlsx" in self.rules:
            to_excel = ToExcel(self.rules["toXlsx"])
            to_excel.write(load_sources[0])
        else:
            return load_sources.pop()
