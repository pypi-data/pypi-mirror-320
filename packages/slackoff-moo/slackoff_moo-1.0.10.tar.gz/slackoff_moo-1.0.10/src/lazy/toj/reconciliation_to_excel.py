from abc import ABC, abstractmethod
from openpyxl import load_workbook
from os.path import join
from re import match, findall
from .reconciliation_exper_processor import NifiExpression, DefaultCalcExpression
from .utils import Utils


class ToWriter(ABC):
    @abstractmethod
    def write(self, data_source):
        pass


# 写入到Excel
class ToExcel(ToWriter):
    var_re = r"([{}\\a-zA-Z0-9\u4e00-\u9fff]+|\[[a-zA-Z0-9\"\'\u4e00-\u9fff{}\\]+])"

    def __init__(self, rules_):
        self.exper_processor = NifiExpression(DefaultCalcExpression())
        if isinstance(rules_, dict):
            self.rules = [rules_]
        elif isinstance(rules_, list):
            self.rules = rules_
        else:
            raise TypeError()

    @staticmethod
    def load_xlsx(rules):
        target = rules["target"]
        if "before" in target:
            Utils.handle(target["before"], "before", None)
        if "dir" not in target:
            raise ValueError("target must contain 'dir'")
        if "filename" not in target:
            raise ValueError("target must contain 'filename'")
        xlsx_path = join(target["dir"], target["filename"])
        if not isinstance(xlsx_path, str):
            raise ValueError("target must contain 'filename'")
        if "sheet" not in target:
            raise ValueError("target must contain 'sheet'")
        sheet_params = target["sheet"]
        if "name" not in sheet_params:
            raise ValueError("target must contain 'name'")
        return [sheet_params, xlsx_path]

    def write(self, data_source):
        for rule in self.rules:
            if "target" not in rule:
                raise ValueError("rules_ must contain 'target'")
            if "writer" not in rule:
                raise ValueError("rules_ must contain 'writer'")
            if data_source is None:
                raise ValueError("数据源为空，请检查数据文件是否有数据")
            if "writer" not in rule:
                raise ValueError("rules must contain 'writer'")
            writer_rules = rule["writer"]
            if not (isinstance(writer_rules, dict) or isinstance(writer_rules, list)):
                raise ValueError()
            if isinstance(writer_rules, dict):
                writer_rules["writer"] = [writer_rules["writer"]]
            excel = self.load_xlsx(rule)
            workbook = load_workbook(excel[1])
            sheet = workbook.get_sheet_by_name(excel[0]["name"])
            for writer_rule in writer_rules:
                self.doWrite(writer_rule, data_source, sheet)
            workbook.save(excel[1])
            workbook.close()

    # 用row确定操作的数据范围，数据会给到col做运算取值
    def doWrite(self, writer_rule, data_source, sheet):
        if "row" not in writer_rule:
            writer_rule["row"] = ""
        row_chain = writer_rule["row"]
        if not isinstance(row_chain, list):
            raise ValueError("row must contain 'list'")
        if "col" not in writer_rule:
            raise ValueError("rules must contain 'col'")
        col = writer_rule["col"]  # 必须是一个字典
        if not isinstance(col, (dict, list)):
            raise ValueError("col must contain 'dict or list'")
        if isinstance(col, dict):
            if not all(match(r"\d+", c) for c in col.keys()):
                raise ValueError("col must contain 'int'")
        start_row_num = 0
        if "start_row_num" in writer_rule:
            start_row_num = writer_rule["start_row_num"]

        # 把解析到的值写入到sheet
        row_num = 1
        if isinstance(data_source, dict):
            for vk, val in data_source.items():  # 字典数据
                row_data = {vk: val}
                is_success = self.cell(writer_rule, col, row_chain, row_data, row_num, start_row_num, sheet)
                if is_success:
                    row_num += 1

        elif isinstance(data_source, list):
            for row_data in data_source:
                is_success = self.cell(writer_rule, col, row_chain, row_data, row_num, start_row_num, sheet)
                if is_success:
                    row_num += 1

    def cell(self, writer_rule, col, row_chain, row_data, row_num, start_row_num, sheet):
        if ("filter" not in writer_rule
                or Utils.handle(writer_rule["filter"], "filter", {"row": row_data})):
            cel_values = self.do_cell(col, row_chain, row_data, row_num, start_row_num)
            [sheet.cell(row=row_num + start_row_num, column=i + 1, value=cel_values[i]) for i in range(len(cel_values))]
            return True
        return False

    # 确定数据范围(row)，解析并计算表达式col_exper
    def do_cell(self, col, row_chain, row_data, row_num, start_row_num):
        if row_chain is not None:
            row_data = self.to_(row_chain, 0, row_data)  # 确定数据区间
        if isinstance(col, dict):
            col_values = []  # 列的所有值
            last_col_index = 1
            col_keys = col.keys()
            for col_index in col_keys:  # 一列一列的匹配
                col_exper = col[col_index]
                col_value = self.do_cell_value(exper=col_exper, row_num=row_num + start_row_num, row_data=row_data)
                while last_col_index < int(col_index):
                    col_values.append(None)
                    last_col_index += 1
                col_values.append(col_value)
                last_col_index += 1
            return col_values
        elif isinstance(col, list):
            col_values = []  # 列的所有值
            for exper in col:
                value = self.do_cell_value(exper, row_num, row_data)
                col_values.append(value)
            return col_values

    def do_cell_value(self, exper, row_num, row_data):
        col_value = None
        if callable(exper):
            col_value = exper(row_num, row_data)
        elif isinstance(exper, str):
            if isinstance(row_data, list):
                col_value = self.exper_processor.processor(exper_=exper, context_={"row": row_data})
            else:
                col_value = self.exper_processor.processor(exper_=exper, context_=row_data)
        elif not (callable(exper) or isinstance(exper, str)):
            raise ValueError("列规则类型只能是字符串和函数")
        return col_value

    # 深度优先搜索，通过row_chain的key去访问数组或字典
    def to_(self, row_chain, i, obj):
        if i >= len(row_chain):
            return obj
        row_data = obj.copy()
        # 确定写入行区间
        target_key = row_chain[i]
        if isinstance(obj, dict):
            each_key_stack = [item for item in row_data.keys()]
            for source_key in each_key_stack:
                if not isinstance(source_key, str):
                    return None
                if match(target_key, source_key) is not None:
                    row_data = row_data[source_key]
                    return self.to_(row_chain, i + 1, row_data)
                return None
        elif isinstance(obj, list):
            index = findall(r"^(\d+)$", target_key)
            if index == 0:
                raise ValueError(target_key, "应该是访问数组，下标应该为数字")
            return obj[int(index[0])]
